from PySide6.QtWidgets import QApplication, QMainWindow, QTableWidgetItem, QFileDialog
from ui_main import Ui_MainWindow
from dataTeamManager import *
from dataResultManager import *
from openpyxl import load_workbook


# 注意 这里选择的父类 要和你UI文件窗体一样的类型
# 主窗口是 QMainWindow， 表单是 QWidget， 对话框是 QDialog
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 使用ui文件导入定义界面类
        self.dataResult = None
        self.resultPath = None
        self.dataTeam = None
        self.ui = Ui_MainWindow()
        # 初始化界面
        self.ui.setupUi(self)
        self.ui.boxCode.currentTextChanged.connect(self.changeCode)
        self.ui.addTeam.clicked.connect(self.addTeam)
        self.ui.delTeam.clicked.connect(self.delTeam)
        self.ui.editTeam.clicked.connect(self.editTeam)
        self.ui.listTeam.clicked.connect(self.clickTeam)
        self.ui.openExcelResult.clicked.connect(self.openExcelResult)
        self.ui.boxRound.currentTextChanged.connect(self.updateResult)
        self.ui.listResult.clicked.connect(self.clickResult)
        self.ui.addResult.clicked.connect(self.addResult)
        self.ui.delResult.clicked.connect(self.delResult)
        self.ui.editResult.clicked.connect(self.editResult)
        self.ui.tableDetail.clicked.connect(self.clickDetail)
        self.ui.addDetailResult.clicked.connect(self.addDetail)
        self.ui.delDetailResult.clicked.connect(self.delDetail)
        self.ui.editDetailResult.clicked.connect(self.editDetail)
        self.ui.importResult.clicked.connect(self.importResult)
        self.ui.exportTLResult.clicked.connect(self.exportTLResult)
        self.ui.addRound.clicked.connect(self.addRound)
        self.ui.delRound.clicked.connect(self.delRound)
        # 初始化数据
        self.initDataTeam()
        self.initDataResult()

    def initDataTeam(self):
        listTeam = self.ui.listTeam
        self.dataTeam = DataTeamManager("dataTeam" + self.ui.boxCode.currentText() + ".json")
        while listTeam.count() > 0:
            listTeam.takeItem(0)
        for name in self.dataTeam.getTeams():
            listTeam.addItem(name)

    def changeCode(self):
        self.initDataTeam()

    def addTeam(self):
        if len(self.ui.textTeam.text()) == 0:
            return
        listTeam = self.ui.listTeam
        if self.dataTeam is not None:
            listTeam.addItem(self.ui.textTeam.text())
            self.dataTeam.addTeam(self.ui.textTeam.text())

    def delTeam(self):
        listTeam = self.ui.listTeam
        if self.dataTeam is not None:
            self.dataTeam.delTeam(listTeam.currentItem().text())
            listTeam.takeItem(listTeam.currentRow())

    def editTeam(self):
        listTeam = self.ui.listTeam
        if self.dataTeam is not None:
            self.dataTeam.editTeam(listTeam.currentItem().text(), self.ui.textTeam.text())
            listTeam.currentItem().setText(self.ui.textTeam.text())

    def clickTeam(self):
        listTeam = self.ui.listTeam
        tablePlayer = self.ui.tablePlayer
        self.ui.textTeam.setText(listTeam.currentItem().text())
        while tablePlayer.rowCount() > 0:
            tablePlayer.removeRow(0)
        for player in self.dataTeam.getPlayers(listTeam.currentItem().text()):
            tablePlayer.insertRow(tablePlayer.rowCount())
            tablePlayer.setItem(tablePlayer.rowCount() - 1, 0, QTableWidgetItem(player["nickname"]))
            tablePlayer.setItem(tablePlayer.rowCount() - 1, 1, QTableWidgetItem(player["id"]))
            tablePlayer.setItem(tablePlayer.rowCount() - 1, 2, QTableWidgetItem(player["race"]))

    def initDataResult(self):
        self.dataResult = DataResultManager("dataResult" + self.ui.boxCode.currentText() + ".json")
        self.updateRound()
        self.updateResult()

    def updateRound(self):
        self.ui.boxRound.clear()
        for roundName in self.dataResult.getRoundList():
            self.ui.boxRound.addItem(roundName)

    def updateResult(self):
        roundName = self.ui.boxRound.currentText()
        listResult = self.ui.listResult
        while listResult.count() > 0:
            listResult.takeItem(0)
        for result in self.dataResult.getResultList(roundName):
            listResult.addItem(result)

    def updateDetail(self):
        tableDetail = self.ui.tableDetail
        while tableDetail.rowCount() > 0:
            tableDetail.removeRow(0)
        for detail in self.dataResult.getDetail(self.ui.boxRound.currentText(), self.ui.listResult.currentRow()):
            tableDetail.insertRow(tableDetail.rowCount())
            tableDetail.setItem(tableDetail.rowCount() - 1, 0, QTableWidgetItem(detail["player1"]))
            tableDetail.setItem(tableDetail.rowCount() - 1, 1, QTableWidgetItem(detail["player2"]))
            tableDetail.setItem(tableDetail.rowCount() - 1, 2, QTableWidgetItem(detail["map"]))
            tableDetail.setItem(tableDetail.rowCount() - 1, 3, QTableWidgetItem(detail["winner"]))

    def addRound(self):
        if len(self.ui.editRound.text()) == 0:
            return
        self.dataResult.addRound(self.ui.editRound.text())
        self.updateRound()

    def delRound(self):
        self.dataResult.delRound(self.ui.boxRound.currentText())
        self.updateRound()

    def clickResult(self):
        result = self.dataResult.getResult(self.ui.boxRound.currentText(), self.ui.listResult.currentRow())
        self.ui.editHomeTeam.setText(result["home"])
        self.ui.editAwayTeam.setText(result["away"])
        self.updateDetail()

    def addResult(self):
        self.dataResult.addResult(self.ui.boxRound.currentText(), self.ui.editHomeTeam.text(),
                                  self.ui.editAwayTeam.text())
        self.updateResult()

    def delResult(self):
        self.dataResult.delResult(self.ui.boxRound.currentText(), self.ui.listResult.currentRow())
        self.updateResult()

    def editResult(self):
        self.dataResult.editResult(self.ui.boxRound.currentText(),
                                   self.ui.listResult.currentRow(),
                                   self.ui.editHomeTeam.text(),
                                   self.ui.editAwayTeam.text())
        self.updateResult()

    def clickDetail(self):
        detail = self.dataResult.getDetail(self.ui.boxRound.currentText(), self.ui.listResult.currentRow())[
            self.ui.tableDetail.currentRow()]
        self.ui.editPlayer1.setText(detail["player1"])
        self.ui.editPlayer2.setText(detail["player2"])
        self.ui.editMap.setText(detail["map"])
        self.ui.editWinner.setText(detail["winner"])

    def addDetail(self):
        self.dataResult.addDetail(self.ui.boxRound.currentText(), self.ui.listResult.currentRow(),
                                  self.ui.editPlayer1.text(), self.ui.editPlayer2.text(), self.ui.editMap.text(),
                                  self.ui.editWinner.text())
        self.updateDetail()

    def delDetail(self):
        self.dataResult.delDetail(self.ui.boxRound.currentText(), self.ui.listResult.currentRow(),
                                  self.ui.tableDetail.currentRow())
        self.updateDetail()

    def editDetail(self):
        self.dataResult.editDetail(self.ui.boxRound.currentText(), self.ui.listResult.currentRow(),
                                   self.ui.tableDetail.currentRow(),
                                   self.ui.editPlayer1.text(), self.ui.editPlayer2.text(), self.ui.editMap.text(),
                                   self.ui.editWinner.text())
        self.updateDetail()

    def openExcelResult(self):
        (file_path, file_type) = QFileDialog.getOpenFileName(self, "导入结果")
        self.resultPath = file_path
        self.excelResult = load_workbook(self.resultPath)
        for name in self.excelResult.sheetnames:
            self.ui.boxSheetResult.addItem(name)

    def importResult(self):
        sheet = self.excelResult[self.ui.boxSheetResult.currentText()]
        curRow = 0
        curCol = 0
        for row in sheet.rows:
            for cell in row:
                if cell.value == self.ui.boxRound.currentText():
                    curRow = cell.row + 2
                    curCol = cell.column
        while sheet.cell(curRow, curCol).value is not None:
            self.dataResult.addResult(self.ui.boxRound.currentText(), sheet.cell(curRow, curCol).value,
                                      sheet.cell(curRow, curCol + 4).value)
            for i in range(1, 7):
                self.dataResult.addDetail(self.ui.boxRound.currentText(), self.ui.listResult.count() - 1,
                                          sheet.cell(curRow + i, curCol).value,
                                          sheet.cell(curRow + i, curCol + 4).value,
                                          sheet.cell(curRow + i, curCol + 2).value,
                                          str(int(sheet.cell(curRow + i, curCol + 3).value) + 1))
            curRow += 8
        self.updateResult()
        self.updateDetail()

    def exportTLResult(self):
        (file_path, file_type) = QFileDialog.getSaveFileName(self, "导出TL代码")
        res = ""
        for i in range(0, self.ui.listResult.count()):
            res += self.dataResult.exportResult(self.ui.boxRound.currentText(), i, self.ui.editDate.text())

        with open(file_path, "w") as f:
            f.write(res)


app = QApplication([])
mainWindows = MainWindow()
mainWindows.show()
app.exec()
